from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from collections import Counter


class ExportadorExcel:
    """Exporta una lista de registros de seguimiento a un archivo Excel."""

    # (clave del diccionario, título de la columna en Excel). El orden de esta
    # lista define el orden de las columnas en la planilla.
    COLUMNAS = [
        ("inicio_dia", "Fecha"),
        ("inicio_hora", "Hora inicio"),
        ("fin_hora", "Hora fin"),
        ("punto", "Ubicación (Access Point)"),
        ("mac_ap", "MAC del AP"),
        ("session_time", "Duración (seg)"),
        ("input_bytes", "Entrada (bytes)"),
        ("output_bytes", "Salida (bytes)"),
        ("mac_cliente", "MAC del dispositivo"),
    ]

    def exportar(self, registros, usuario, desde, hasta, repositorio, ruta_salida):
        """Crea y guarda el archivo .xlsx con el seguimiento del usuario."""
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Seguimiento"

        # --- Fila de título (información del seguimiento) ---
        hoja.append([f"Seguimiento del usuario: {usuario}"])
        hoja.append([f"Rango de fechas: {desde} a {hasta}"])
        hoja.append([f"Total de conexiones: {len(registros)}"])
        hoja.append([])   # fila en blanco de separación

        # --- Fila de encabezados (en negrita) ---
        fila_titulos = [titulo for _, titulo in self.COLUMNAS]
        hoja.append(fila_titulos)
        nro_fila_titulos = hoja.max_row
        for celda in hoja[nro_fila_titulos]:
            celda.font = Font(bold=True)
            celda.alignment = Alignment(horizontal="center")

        # --- Filas de datos ---
        for r in registros:
            # agregamos al diccionario el nombre corto del punto (Access Point)
            r_punto = dict(r)
            r_punto["punto"] = repositorio.etiqueta(r["mac_ap"])
            hoja.append([r_punto[clave] for clave, _ in self.COLUMNAS])

        # --- Ajuste de ancho de columnas para que se lea cómodo ---
        anchos = [12, 12, 12, 18, 24, 14, 16, 16, 22]
        for indice, ancho in enumerate(anchos, start=1):
            letra = hoja.cell(row=1, column=indice).column_letter
            hoja.column_dimensions[letra].width = ancho

        libro.save(ruta_salida)
        return ruta_salida

    # -----------------------------------------------------------------
    # Exportación de registros INVÁLIDOS (descartados) para analizarlos
    # -----------------------------------------------------------------

    def _campo_que_fallo(self, motivo):
        """Saca del texto del motivo el nombre del campo que falló."""
        if motivo.startswith("campo "):
            return motivo.split()[1]      # la palabra que sigue a "campo"
        return "fila incompleta"

    def exportar_invalidos(self, descartados, cabecera, ruta_salida):
        """Exporta TODOS los registros descartados a un Excel con dos hojas:"""
        libro = Workbook()

        # ---------- Hoja 1: Resumen por campo ----------
        hoja_resumen = libro.active
        hoja_resumen.title = "Resumen"

        total = len(descartados)
        hoja_resumen.append(["Resumen de registros descartados"])
        hoja_resumen.append([f"Total descartados: {total}"])
        hoja_resumen.append([])

        titulos_resumen = ["Campo que falló", "Cantidad", "Porcentaje"]
        hoja_resumen.append(titulos_resumen)
        for celda in hoja_resumen[hoja_resumen.max_row]:
            celda.font = Font(bold=True)
            celda.alignment = Alignment(horizontal="center")

        # Contamos cuántos descartes hubo por cada campo que falló.
        conteo = Counter(self._campo_que_fallo(m) for _, m, _ in descartados)
        for campo, cantidad in conteo.most_common():
            porcentaje = f"{cantidad / total * 100:.1f}%" if total else "0%"
            hoja_resumen.append([campo, cantidad, porcentaje])

        for letra, ancho in zip("ABC", [22, 12, 12]):
            hoja_resumen.column_dimensions[letra].width = ancho

        # ---------- Hoja 2: Detalle fila por fila ----------
        hoja_detalle = libro.create_sheet("Detalle")

        # Encabezados: Línea, Motivo y luego las columnas originales del CSV.
        # Si algún título de columna viene vacío (las comas de más al final),
        # lo reemplazamos por un nombre genérico para que se entienda.
        columnas_csv = [
            (titulo if titulo.strip() != "" else f"col_extra_{i}")
            for i, titulo in enumerate(cabecera, start=1)
        ]
        titulos_detalle = ["Línea CSV", "Motivo del descarte"] + columnas_csv
        hoja_detalle.append(titulos_detalle)
        for celda in hoja_detalle[hoja_detalle.max_row]:
            celda.font = Font(bold=True)
            celda.alignment = Alignment(horizontal="center")

        # Una fila por cada registro descartado, con sus columnas tal cual
        # venían en el CSV (así se ve si las columnas estaban corridas).
        for numero_linea, motivo, fila in descartados:
            hoja_detalle.append([numero_linea, motivo] + list(fila))

        # Ancho cómodo para las primeras columnas (las más importantes).
        anchos_detalle = [10, 42, 8, 12, 16, 16, 16, 14, 14, 12, 12, 14, 14, 14, 24, 20, 28]
        for indice, ancho in enumerate(anchos_detalle, start=1):
            letra = hoja_detalle.cell(row=1, column=indice).column_letter
            hoja_detalle.column_dimensions[letra].width = ancho

        libro.save(ruta_salida)
        return ruta_salida